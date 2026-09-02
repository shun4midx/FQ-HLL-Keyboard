############################################
# Copyright (c) 2026 Shun/修海 (@shun4midx) #
# Project: FQ-HLL-Keyboard                 #
# File Type: Python file                   #
# File: extract_tw_chars.py                #
############################################

import json
from collections import defaultdict
from openpyxl import load_workbook

INPUT = "附錄1、民國112年語料字頻表.xlsx"
OUTPUT = "taiwan_char_freq.json"

wb = load_workbook(INPUT, read_only=True, data_only=True)
ws = wb["112年語料字頻表"]

freq = defaultdict(int)
duplicates = defaultdict(list)

for row in ws.iter_rows(min_row=2, values_only=True):
    char = row[1]
    count = row[2]

    if not char or count is None:
        continue

    char = str(char).strip()

    # Only actual single-codepoint characters
    if len(char) != 1:
        continue

    if isinstance(count, str):
        count = count.replace(",", "").strip()

    count = int(count)

    if char in freq:
        duplicates[char].append(count)

    freq[char] += count

# Optional: sort by final total frequency descending.
freq = dict(sorted(freq.items(), key=lambda item: item[1], reverse=True))

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump(freq, f, ensure_ascii=False, separators=(",", ":"))

print(f"saved {len(freq)} unique characters")
print(f"{len(duplicates)} characters occurred more than once")

for char, extra_counts in list(duplicates.items())[:20]:
    print(repr(char), "total =", freq[char], "additional rows =", extra_counts)