############################################
# Copyright (c) 2026 Shun/修海 (@shun4midx) #
# Project: FQ-HLL-Keyboard                 #
# File Type: Python file                   #
# File: extract_tw_words.py                #
############################################

from openpyxl import load_workbook
import json

INPUT = "附錄2、民國112年語料詞頻表.xlsx"
OUTPUT = "taiwan_word_freq.json"

SHEETS = [
    "112年語料詞頻表(多音節詞)-1",
    "112年語料詞頻表(多音節詞)-2",
    "112年語料詞頻表(多音節詞)-3",
    "112年語料詞頻表(多音節詞)-4",
]

MIN_FREQ = 5

print("Opening workbook...")
wb = load_workbook(INPUT, read_only=True, data_only=True)

word_freq = {}

for sheet_name in SHEETS:
    print(f"Reading {sheet_name}...")
    ws = wb[sheet_name]

    first = True
    count = 0

    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue

        rank = row[0]
        word = row[1]
        freq = row[2]

        if word is None or freq is None:
            continue

        try:
            freq = int(freq)
        except (TypeError, ValueError):
            continue

        if freq < MIN_FREQ:
            break

        word = str(word).strip()
        if not word:
            continue
        
        old = word_freq.get(word)
        if old is None or freq > old:
            word_freq[word] = freq

        count += 1

    print(f"  kept {count:,} rows")

wb.close()

print(f"\nWriting {len(word_freq):,} unique words...")
with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump(word_freq, f, ensure_ascii=False, separators=(",", ":"))

print(f"Done: {OUTPUT}")